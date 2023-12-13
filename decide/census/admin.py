from django.contrib import admin

from .models import Census, CensusYesNo, CensusByPreference
from django.contrib import messages
from django.contrib.admin.helpers import ActionForm
from django import forms
from django.http import HttpResponse
from openpyxl import Workbook


class ReuseActionForm(ActionForm):
    id_to_reuse = forms.IntegerField(required=False)
    id_to_reuse.label = "ID de la votación (Reutilizar):"


class CensusAdmin(admin.ModelAdmin):
    list_display = ("voting_id", "voter_id", "group")
    list_filter = ("voting_id", "group")
    search_fields = ("voter_id",)

    def exportar_a_excel(modeladmin, request, queryset):
        workbook = Workbook()
        sheet = workbook.active

        sheet.append(
            ["ID Votacion", "ID Votante", "Grupo"]
        )  # El append funciona en filas, de izquierda a derecha

        for elemento in queryset:
            sheet.append([elemento.voting_id, elemento.voter_id, elemento.group])

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = "attachment; filename=exportacion_censo.xlsx"
        workbook.save(response)

        return response

    exportar_a_excel.short_description = "Exportar a Excel"

    def reuse_action(modeladmin, request, queryset):
        reuse_voting_id = request.POST.get("id_to_reuse")

        if reuse_voting_id is not None and reuse_voting_id.strip():
            modeladmin.message_user(request, f"ID introducido: {reuse_voting_id}")

            for census in queryset.all():
                if Census.objects.filter(
                    voting_id=reuse_voting_id, voter_id=census.voter_id
                ).exists():
                    messages.error(
                        request,
                        f"Ya existe Censo con voter_id {census.voter_id} y voting_id {reuse_voting_id} en la base de datos.",
                    )
                    continue  # Salta al siguiente censo en lugar de intentar guardarlo
                re_census = Census()
                re_census.voter_id = census.voter_id
                re_census.voting_id = reuse_voting_id
                re_census.save()
        else:
            messages.error(
                request,
                "Error: Formulario no válido. Asegúrate de ingresar un ID válido.",
            )

    reuse_action.short_description = "Reutilizar Censo"

    actions = [reuse_action, exportar_a_excel]
    action_form = ReuseActionForm


class CensusByPreferenceAdmin(admin.ModelAdmin):
    list_display = ("voting_id", "voter_id", "group")
    list_filter = ("voting_id","group")
    search_fields = ("voter_id",)

    def exportar_a_excel(modeladmin, request, queryset):
        workbook = Workbook()
        sheet = workbook.active

        sheet.append(
            ["ID Votacion", "ID Votante"]
        )  # El append funciona en filas, de izquierda a derecha

        for elemento in queryset:
            sheet.append([elemento.voting_id, elemento.voter_id])

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = "attachment; filename=exportacion_censo.xlsx"
        workbook.save(response)

        return response

    exportar_a_excel.short_description = "Exportar a Excel"

    def reuse_action(modeladmin, request, queryset):
        reuse_voting_id = request.POST.get("id_to_reuse")

        if reuse_voting_id is not None and reuse_voting_id.strip():
            modeladmin.message_user(request, f"ID introducido: {reuse_voting_id}")

            for census in queryset.all():
                if CensusByPreference.objects.filter(
                    voting_id=reuse_voting_id, voter_id=census.voter_id
                ).exists():
                    messages.error(
                        request,
                        f"Ya existe Censo con voter_id {census.voter_id} y voting_id {reuse_voting_id} en la base de datos.",
                    )
                    continue  # Salta al siguiente censo en lugar de intentar guardarlo
                re_census = CensusByPreference()
                re_census.voter_id = census.voter_id
                re_census.voting_id = reuse_voting_id
                re_census.save()
        else:
            messages.error(
                request,
                "Error: Formulario no válido. Asegúrate de ingresar un ID válido.",
            )

    reuse_action.short_description = "Reutilizar Censo por preferencia"

    actions = [reuse_action, exportar_a_excel]
    action_form = ReuseActionForm
    
class CensusYesNoAdmin(admin.ModelAdmin):
    list_display = ("voting_id", "voter_id", "group")
    list_filter = ("voting_id", "group")
    search_fields = ("voter_id",)

    def exportar_a_excel(modeladmin, request, queryset):
        workbook = Workbook()
        sheet = workbook.active

        sheet.append(
            ["ID Votacion", "ID Votante"]
        )  # El append funciona en filas, de izquierda a derecha

        for elemento in queryset:
            sheet.append([elemento.voting_id, elemento.voter_id])

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = "attachment; filename=exportacion_censo.xlsx"
        workbook.save(response)

        return response

    exportar_a_excel.short_description = "Exportar a Excel"

    def reuse_action(modeladmin, request, queryset):
        reuse_voting_id = request.POST.get("id_to_reuse")

        if reuse_voting_id is not None and reuse_voting_id.strip():
            modeladmin.message_user(request, f"ID introducido: {reuse_voting_id}")

            for census in queryset.all():
                if CensusYesNo.objects.filter(
                    voting_id=reuse_voting_id, voter_id=census.voter_id
                ).exists():
                    messages.error(
                        request,
                        f"Ya existe Censo con voter_id {census.voter_id} y voting_id {reuse_voting_id} en la base de datos.",
                    )
                    continue  # Salta al siguiente censo en lugar de intentar guardarlo
                re_census = CensusYesNo()
                re_census.voter_id = census.voter_id
                re_census.voting_id = reuse_voting_id
                re_census.save()
        else:
            messages.error(
                request,
                "Error: Formulario no válido. Asegúrate de ingresar un ID válido.",
            )

    reuse_action.short_description = "Reutilizar Censo"

    actions = [reuse_action, exportar_a_excel]
    action_form = ReuseActionForm

admin.site.register(Census, CensusAdmin)
admin.site.register(CensusByPreference, CensusByPreferenceAdmin)
admin.site.register(CensusYesNo, CensusYesNoAdmin)


