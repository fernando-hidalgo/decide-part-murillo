from django.db.utils import IntegrityError
from django.core.exceptions import ObjectDoesNotExist
from rest_framework import generics
from rest_framework.response import Response
from rest_framework.generics import CreateAPIView
from rest_framework.status import (
    HTTP_201_CREATED as ST_201,
    HTTP_204_NO_CONTENT as ST_204,
    HTTP_400_BAD_REQUEST as ST_400,
    HTTP_401_UNAUTHORIZED as ST_401,
    HTTP_409_CONFLICT as ST_409,
)

from base.perms import UserIsStaff
from .models import Census, CensusByPreference, CensusYesNo, CensusMultiChoice
from django.contrib import messages
import openpyxl
from django.http import HttpResponseRedirect
from django.views.generic.base import TemplateView
from .serializers import CensusSerializer
from rest_framework.permissions import AllowAny
from django.views.decorators.csrf import csrf_exempt


class CensusCreate(generics.ListCreateAPIView):
    permission_classes = (UserIsStaff,)

    @csrf_exempt
    def create(self, request, *args, **kwargs):
        voting_id = request.data.get("voting_id")
        voters = request.data.get("voters")
        try:
            for voter in voters:
                census = Census(voting_id=voting_id, voter_id=voter)
                census.save()
        except IntegrityError:
            return Response("Error try to create census", status=ST_409)
        return Response("Census created", status=ST_201)

    @csrf_exempt
    def list(self, request, *args, **kwargs):
        voting_id = request.GET.get("voting_id")
        voters = Census.objects.filter(voting_id=voting_id).values_list(
            "voter_id", flat=True
        )
        return Response({"voters": voters})


class CensusCreateView(CreateAPIView):
    queryset = Census.objects.all()
    serializer_class = CensusSerializer


class CensusListCreateAPIView(generics.ListCreateAPIView):
    queryset = Census.objects.all()
    serializer_class = CensusSerializer
    permission_classes = [AllowAny]
    # Si necesitas permisos específicos, puedes añadirlos aquí


class CensusDetail(generics.RetrieveDestroyAPIView):
    def destroy(self, request, voting_id, *args, **kwargs):
        voters = request.data.get("voters")
        census = Census.objects.filter(voting_id=voting_id, voter_id__in=voters)
        census.delete()
        return Response("Voters deleted from census", status=ST_204)

    def retrieve(self, request, voting_id, *args, **kwargs):
        voter = request.GET.get("voter_id")
        try:
            Census.objects.get(voting_id=voting_id, voter_id=voter)
        except ObjectDoesNotExist:
            return Response("Invalid voter", status=ST_401)
        return Response("Valid voter")


class CensusAdminView(TemplateView):
    template_name = "census/admin_census.html"

    def get_context_data(self, **kwargs):
        # Aquí puedes añadir lógica para enviar datos adicionales al template si es necesario
        context = super().get_context_data(**kwargs)
        # Ejemplo: Añadir alguna configuración o dato relevante
        context["extra_info"] = "Algún valor o configuración"
        return context

    def list(self, request, *args, **kwargs):
        voting_id = request.GET.get("voting_id")
        voters = Census.objects.filter(voting_id=voting_id).values_list(
            "voter_id", flat=True
        )
        return Response({"voters": voters})


class CensusImportView(TemplateView):
    template_name = "census/import_census.html"

    def post(self, request, *args, **kwargs):
        if request.method == "POST" and request.FILES["census_file"]:
            census_file = request.FILES["census_file"]
            workbook = openpyxl.load_workbook(census_file)
            sheet = workbook.active

            for row in sheet.iter_rows(min_row=2, values_only=True):
                voting_id = row[0]
                voter_id = row[1]
                group = row[2]

                # Comprobar si ya existe un objeto con la misma pareja de voting_id y voter_id
                existing_census = Census.objects.filter(
                    voting_id=voting_id, voter_id=voter_id
                ).first()

                if not existing_census:
                    Census.objects.create(voting_id=voting_id, voter_id=voter_id, group=group)
                else:
                    messages.error(
                        request,
                        f"Ya existe un registro para la pareja de voting_id={voting_id}, voter_id={voter_id} y group={group}",
                    )

            messages.success(request, "Importación finalizada")
            return HttpResponseRedirect("/census/import/")


class CensusByPreferenceCreate(generics.ListCreateAPIView):
    permission_classes = (UserIsStaff,)

    def create(self, request, *args, **kwargs):
        voting_id = request.data.get("voting_id")
        voters = request.data.get("voters")
        try:
            for voter in voters:
                census = CensusByPreference(voting_id=voting_id, voter_id=voter)
                census.save()
        except IntegrityError:
            return Response("Error try to create census", status=ST_409)
        return Response("Census by preference created", status=ST_201)

    def list(self, request, *args, **kwargs):
        voting_id = request.GET.get("voting_id")
        voters = CensusByPreference.objects.filter(voting_id=voting_id).values_list(
            "voter_id", flat=True
        )
        return Response({"voters": voters})


class CensusByPreferenceDetail(generics.RetrieveDestroyAPIView):
    def destroy(self, request, voting_id, *args, **kwargs):
        voters = request.data.get("voters")
        census = CensusByPreference.objects.filter(
            voting_id=voting_id, voter_id__in=voters
        )
        census.delete()
        return Response("Voters deleted from census by preference", status=ST_204)

    def retrieve(self, request, voting_id, *args, **kwargs):
        voter = request.GET.get("voter_id")
        try:
            CensusByPreference.objects.get(voting_id=voting_id, voter_id=voter)
        except ObjectDoesNotExist:
            return Response("Invalid voter", status=ST_401)
        return Response("Valid voter")


class CensusByPreferenceImportView(TemplateView):
    template_name = "census/import_census.html"

    def post(self, request, *args, **kwargs):
        if request.method == "POST" and request.FILES["census_file"]:
            census_file = request.FILES["census_file"]
            workbook = openpyxl.load_workbook(census_file)
            sheet = workbook.active

            for row in sheet.iter_rows(min_row=2, values_only=True):
                voting_id = row[0]
                voter_id = row[1]
                group = row[2]

                # Comprobar si ya existe un objeto con la misma pareja de voting_id y voter_id
                existing_census = CensusByPreference.objects.filter(
                    voting_id=voting_id, voter_id=voter_id, group=group
                ).first()

                if not existing_census:
                    CensusByPreference.objects.create(
                        voting_id=voting_id, voter_id=voter_id
                    )
                else:
                    messages.error(
                        request,
                        f"Ya existe un registro para la pareja de voting_id={voting_id}, voter_id={voter_id} y grupo={group}",
                    )

            messages.success(request, "Importación finalizada")
            return HttpResponseRedirect("/census/bypreference/import/")
          
class CensusYesNoCreate(generics.ListCreateAPIView):
    permission_classes = (UserIsStaff,)

    def create(self, request, *args, **kwargs):
        voting_id = request.data.get("voting_id")
        voters = request.data.get("voters")
        try:
            for voter in voters:
                census = CensusYesNo(voting_id=voting_id, voter_id=voter)
                census.save()
        except IntegrityError:
            return Response("Error try to create census", status=ST_409)
        return Response("Census created", status=ST_201)

    def list(self, request, *args, **kwargs):
        voting_id = request.GET.get("voting_id")
        voters = CensusYesNo.objects.filter(voting_id=voting_id).values_list(
            "voter_id", flat=True
        )
        return Response({"voters": voters})


class CensusYesNoDetail(generics.RetrieveDestroyAPIView):
    def destroy(self, request, voting_id, *args, **kwargs):
        voters = request.data.get("voters")
        census = CensusYesNo.objects.filter(voting_id=voting_id, voter_id__in=voters)
        census.delete()
        return Response("Voters deleted from census", status=ST_204)

    def retrieve(self, request, voting_id, *args, **kwargs):
        voter = request.GET.get("voter_id")
        try:
            CensusYesNo.objects.get(voting_id=voting_id, voter_id=voter)
        except ObjectDoesNotExist:
            return Response("Invalid voter", status=ST_401)
        return Response("Valid voter")


class CensusYesNoImportView(TemplateView):
    template_name = "census/import_census.html"

    def post(self, request, *args, **kwargs):
        if request.method == "POST" and request.FILES["census_file"]:
            census_file = request.FILES["census_file"]
            workbook = openpyxl.load_workbook(census_file)
            sheet = workbook.active

            for row in sheet.iter_rows(min_row=2, values_only=True):
                voting_id = row[0]
                voter_id = row[1]

                # Comprobar si ya existe un objeto con la misma pareja de voting_id y voter_id
                existing_census = CensusYesNo.objects.filter(
                    voting_id=voting_id, voter_id=voter_id
                ).first()

                if not existing_census:
                    CensusYesNo.objects.create(voting_id=voting_id, voter_id=voter_id)
                else:
                    messages.error(
                        request,
                        f"Ya existe un registro para la pareja de voting_id={voting_id} y voter_id={voter_id}",
                    )

            messages.success(request, "Importación finalizada")
            return HttpResponseRedirect("/census/import/")

class CensusMultiChoiceCreate(generics.ListCreateAPIView):
    permission_classes = (UserIsStaff,)

    def create(self, request, *args, **kwargs):
        voting_id = request.data.get("voting_id")
        voters = request.data.get("voters")
        try:
            for voter in voters:
                census = CensusMultiChoice(voting_id=voting_id, voter_id=voter)
                census.save()
        except IntegrityError:
            return Response("Error try to create census", status=ST_409)
        return Response("Census of multiple choice created", status=ST_201)

    def list(self, request, *args, **kwargs):
        voting_id = request.GET.get("voting_id")
        voters = CensusMultiChoice.objects.filter(voting_id=voting_id).values_list(
            "voter_id", flat=True
        )
        return Response({"voters": voters})


class CensusMultiChoiceDetail(generics.RetrieveDestroyAPIView):
    def destroy(self, request, voting_id, *args, **kwargs):
        voters = request.data.get("voters")
        census = CensusMultiChoice.objects.filter(
            voting_id=voting_id, voter_id__in=voters
        )
        census.delete()
        return Response("Voters deleted from census by multiple choice", status=ST_204)

    def retrieve(self, request, voting_id, *args, **kwargs):
        voter = request.GET.get("voter_id")
        try:
            CensusMultiChoice.objects.get(voting_id=voting_id, voter_id=voter)
        except ObjectDoesNotExist:
            return Response("Invalid voter", status=ST_401)
        return Response("Valid voter")


class CensusMultiChoiceImportView(TemplateView):
    template_name = "census/import_census.html"

    def post(self, request, *args, **kwargs):
        if request.method == "POST" and request.FILES["census_file"]:
            census_file = request.FILES["census_file"]
            workbook = openpyxl.load_workbook(census_file)
            sheet = workbook.active

            for row in sheet.iter_rows(min_row=2, values_only=True):
                voting_id = row[0]
                voter_id = row[1]
                group = row[2]

                # Comprobar si ya existe un objeto con la misma pareja de voting_id y voter_id
                existing_census = CensusMultiChoice.objects.filter(
                    voting_id=voting_id, voter_id=voter_id, group=group
                ).first()

                if not existing_census:
                    CensusMultiChoice.objects.create(
                        voting_id=voting_id, voter_id=voter_id, group=group
                    )
                else:
                    messages.error(
                        request,
                        f"Ya existe un registro para la pareja de voting_id={voting_id}, voter_id={voter_id} y grupo={group}",
                    )

            messages.success(request, "Importación finalizada")
            return HttpResponseRedirect("/census/multichoice/import/")