import secrets
from django.test import TestCase
from django.contrib.staticfiles.testing import StaticLiveServerTestCase

from selenium import webdriver
from selenium.webdriver.common.by import By

from .admin import CensusAdmin, CensusByPreferenceAdmin, CensusYesNoAdmin, CensusMultiChoiceAdmin
from django.http import HttpRequest

from .models import Census, CensusByPreference, CensusYesNo, CensusMultiChoice
from base.tests import BaseTestCase
from datetime import datetime

from voting.models import (
    Voting,
    Question,
    QuestionOption,
    VotingByPreference,
    QuestionByPreference,
    QuestionOptionByPreference,
    VotingYesNo,
    QuestionYesNo,
    VotingMultiChoice,
    QuestionMultiChoice,
    QuestionOptionMultiChoice,
)
from base.models import Auth
from django.conf import settings
from openpyxl import Workbook, load_workbook
from django.core.files.uploadedfile import SimpleUploadedFile
from io import BytesIO
from django.urls import reverse

from django.contrib.admin.sites import AdminSite
from django.contrib import messages

# class CensusFrontendTest(StaticLiveServerTestCase):
#     def setUp(self):
#         self.driver = webdriver.Chrome()
#         self.admin_user = User.objects.create_superuser(
#             username='admin',
#             email='admin@test.com',
#             password='secret'
#         )

#         # Crear una pregunta y opciones para la votación
#         question = Question.objects.create(desc="Descripción de la pregunta")
#         QuestionOption.objects.create(question=question, option="Opción 1")
#         QuestionOption.objects.create(question=question, option="Opción 2")

#         # Crear una votación
#         self.voting = Voting.objects.create(
#             name="Votación de Prueba",
#             desc="Descripción de la Votación",
#             question=question
#         )

#         # (Opcional) Crear otros usuarios para añadir al censo
#         self.user1 = User.objects.create_user(username='user1', password='password1', email='user1@test.com')
#         self.user2 = User.objects.create_user(username='user2', password='password2', email= 'user2@gmail.com')
#         # ... crea más usuarios según sea necesario

#     def test_admin_login_and_access_census(self):
#         self.driver.get(self.live_server_url + '/admin')  # URL de inicio de sesión del admin
#         username_input = self.driver.find_element(By.NAME, 'username')
#         password_input = self.driver.find_element(By.NAME, 'password')
#         username_input.send_keys('admin')
#         password_input.send_keys('secret')
#         time.sleep(30)
#         login_button = self.driver.find_element(By.CSS_SELECTOR, "form input[type='submit']")
#         login_button.click()

#         wait = WebDriverWait(self.driver, 10)  # Define un tiempo máximo de espera de 10 segundos
#         # # Esperar a que el botón de inicio de sesión esté presente y hacer clic
#         # login_button = wait.until(
#         #     EC.presence_of_element_located((By.XPATH, '//input[@type="submit" and @value="Iniciar sesión"]'))
#         # )
#         # login_button.click()

#         url = self.live_server_url + '/census/admin/'

#         # Agrega las credenciales en la URL

#         self.driver.get(url)

#         # Esperar a que se cargue la página de administración de censos

#         # Buscar y rellenar los campos del formulario
#         voting_id_input = wait.until(EC.presence_of_element_located((By.ID, 'id_voting_id')))
#         voter_id_input = wait.until(EC.presence_of_element_located((By.ID, 'id_voter_id')))
#         voting_id_input.send_keys('1')
#         voter_id_input.send_keys('1')
#         submit_button = wait.until(EC.presence_of_element_located((By.ID, 'submit_button')))
#         submit_button.click()

#     def tearDown(self):
#         # Limpiar los datos
#         self.driver.quit()
#         User.objects.all().delete()
#         Question.objects.all().delete()
#         Voting.objects.all().delete()
#         Census.objects.all().delete()


class CensusTestCase(BaseTestCase):
    def setUp(self):
        super().setUp()
        self.census = Census(voting_id=1, voter_id=1)
        self.census.save()

    def tearDown(self):
        super().tearDown()
        self.census = None

    def test_check_vote_permissions(self):
        response = self.client.get(
            "/census/{}/?voter_id={}".format(1, 2), format="json"
        )
        self.assertEqual(response.status_code, 401)
        self.assertEqual(response.json(), "Invalid voter")

        response = self.client.get(
            "/census/{}/?voter_id={}".format(1, 1), format="json"
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json(), "Valid voter")

    def test_list_voting(self):
        response = self.client.get("/census/?voting_id={}".format(1), format="json")
        self.assertEqual(response.status_code, 401)

        self.login(user="noadmin")
        response = self.client.get("/census/?voting_id={}".format(1), format="json")
        self.assertEqual(response.status_code, 403)

        self.login()
        response = self.client.get("/census/?voting_id={}".format(1), format="json")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json(), {"voters": [1]})

    def test_add_new_voters_conflict(self):
        data = {"voting_id": 1, "voters": [1]}
        response = self.client.post("/census/", data, format="json")
        self.assertEqual(response.status_code, 401)

        self.login(user="noadmin")
        response = self.client.post("/census/", data, format="json")
        self.assertEqual(response.status_code, 403)

        self.login()
        response = self.client.post("/census/", data, format="json")
        self.assertEqual(response.status_code, 409)

    def test_add_new_voters(self):
        data = {"voting_id": 2, "voters": [1, 2, 3, 4]}
        response = self.client.post("/census/", data, format="json")
        self.assertEqual(response.status_code, 401)

        self.login(user="noadmin")
        response = self.client.post("/census/", data, format="json")
        self.assertEqual(response.status_code, 403)

        self.login()
        response = self.client.post("/census/", data, format="json")
        self.assertEqual(response.status_code, 201)
        self.assertEqual(len(data.get("voters")), Census.objects.count() - 1)

    def test_destroy_voter(self):
        data = {"voters": [1]}
        response = self.client.delete("/census/{}/".format(1), data, format="json")
        self.assertEqual(response.status_code, 204)
        self.assertEqual(0, Census.objects.count())


class CensusTest(StaticLiveServerTestCase):
    def setUp(self):
        # Load base test functionality for decide
        self.base = BaseTestCase()
        self.base.setUp()

        options = webdriver.ChromeOptions()
        options.headless = True
        self.driver = webdriver.Chrome(options=options)

        super().setUp()

    def tearDown(self):
        super().tearDown()
        self.driver.quit()

        self.base.tearDown()

    def createCensusSuccess(self):
        self.cleaner.get(self.live_server_url + "/admin/login/?next=/admin/")
        self.cleaner.set_window_size(1280, 720)

        self.cleaner.find_element(By.ID, "id_username").click()
        self.cleaner.find_element(By.ID, "id_username").send_keys("decide")

        self.cleaner.find_element(By.ID, "id_password").click()
        self.cleaner.find_element(By.ID, "id_password").send_keys("decide")

        self.cleaner.find_element(By.ID, "id_password").send_keys("Keys.ENTER")

        self.cleaner.get(self.live_server_url + "/admin/census/census/add")
        now = datetime.now()
        self.cleaner.find_element(By.ID, "id_voting_id").click()
        self.cleaner.find_element(By.ID, "id_voting_id").send_keys(
            now.strftime("%m%d%M%S")
        )
        self.cleaner.find_element(By.ID, "id_voter_id").click()
        self.cleaner.find_element(By.ID, "id_voter_id").send_keys(
            now.strftime("%m%d%M%S")
        )
        self.cleaner.find_element(By.NAME, "_save").click()

        self.assertTrue(
            self.cleaner.current_url == self.live_server_url + "/admin/census/census"
        )

    def createCensusEmptyError(self):
        self.cleaner.get(self.live_server_url + "/admin/login/?next=/admin/")
        self.cleaner.set_window_size(1280, 720)

        self.cleaner.find_element(By.ID, "id_username").click()
        self.cleaner.find_element(By.ID, "id_username").send_keys("decide")

        self.cleaner.find_element(By.ID, "id_password").click()
        self.cleaner.find_element(By.ID, "id_password").send_keys("decide")

        self.cleaner.find_element(By.ID, "id_password").send_keys("Keys.ENTER")

        self.cleaner.get(self.live_server_url + "/admin/census/census/add")

        self.cleaner.find_element(By.NAME, "_save").click()

        self.assertTrue(
            self.cleaner.find_element_by_xpath(
                "/html/body/div/div[3]/div/div[1]/div/form/div/p"
            ).text
            == "Please correct the errors below."
        )
        self.assertTrue(
            self.cleaner.current_url
            == self.live_server_url + "/admin/census/census/add"
        )

    def createCensusValueError(self):
        self.cleaner.get(self.live_server_url + "/admin/login/?next=/admin/")
        self.cleaner.set_window_size(1280, 720)

        self.cleaner.find_element(By.ID, "id_username").click()
        self.cleaner.find_element(By.ID, "id_username").send_keys("decide")

        self.cleaner.find_element(By.ID, "id_password").click()
        self.cleaner.find_element(By.ID, "id_password").send_keys("decide")

        self.cleaner.find_element(By.ID, "id_password").send_keys("Keys.ENTER")

        self.cleaner.get(self.live_server_url + "/admin/census/census/add")
        now = datetime.now()
        self.cleaner.find_element(By.ID, "id_voting_id").click()
        self.cleaner.find_element(By.ID, "id_voting_id").send_keys("64654654654654")
        self.cleaner.find_element(By.ID, "id_voter_id").click()
        self.cleaner.find_element(By.ID, "id_voter_id").send_keys("64654654654654")
        self.cleaner.find_element(By.NAME, "_save").click()

        self.assertTrue(
            self.cleaner.find_element_by_xpath(
                "/html/body/div/div[3]/div/div[1]/div/form/div/p"
            ).text
            == "Please correct the errors below."
        )
        self.assertTrue(
            self.cleaner.current_url
            == self.live_server_url + "/admin/census/census/add"
        )


class CensusImportViewTest(BaseTestCase):
    def setUp(self):
        super().setUp()

    def create_voting(self):
        q = Question(desc="test_question")
        q.save()

        options = [
            QuestionOption(question=q, option=f"option {i + 1}") for i in range(3)
        ]
        QuestionOption.objects.bulk_create(options)

        v = Voting(name="test_voting", question=q)
        v.save()

        auth, _ = Auth.objects.get_or_create(
            url=settings.BASEURL, defaults={"me": True, "name": "test_auth"}
        )
        v.auths.add(auth)

        return v

    def test_census_import_view(self):
        self.create_voting()
        test_group="Test Group"

        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Voting ID", "Voter ID", "Group"])
        sheet.append([1, 1, test_group])
        sheet.append([1, 2, test_group])
        sheet.append([1, 1, test_group])  # Censo repetido, dará error

        file_buffer = BytesIO()
        workbook.save(file_buffer)
        file_buffer.seek(0)

        excel_file = SimpleUploadedFile("census.xlsx", file_buffer.read())

        url = reverse("import_census")

        response = self.client.post(url, {"census_file": excel_file}, follow=True)

        self.assertEqual(response.status_code, 200)

        census_data = Census.objects.all()
        self.assertEqual(census_data.count(), 2)
        self.assertEqual(census_data[0].voting_id, 1)
        self.assertEqual(census_data[0].voter_id, 1)
        self.assertEqual(census_data[1].voting_id, 1)
        self.assertEqual(census_data[1].voter_id, 2)

        messages = list(response.context["messages"])
        expected_messages = [
            "Ya existe un registro para la pareja de voting_id=1, voter_id=1 y group=Test Group",
            "Importación finalizada",
        ]
        self.assertEqual([str(msg) for msg in messages], expected_messages)


class AdminExportToExcelTest(TestCase):
    def setUp(self):
        self.site = AdminSite()
        self.model_admin = CensusAdmin(model=Census, admin_site=self.site)
        self.census = Census.objects.create(voter_id=1, voting_id=1)

    def test_export_to_excel(self):
        request = HttpRequest()
        request.method = "POST"

        # Simular la selección de objetos en el panel de administración
        queryset = Census.objects.filter(pk=self.census.pk)

        # Ejecutar la acción de exportar a Excel
        response = self.model_admin.exportar_a_excel(request, queryset)

        # Verificar que la respuesta es un archivo Excel
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertTrue(
            "attachment; filename=exportacion_censo.xlsx"
            in response["Content-Disposition"]
        )

        # Cargar el libro de trabajo desde el contenido de la respuesta
        content = BytesIO(response.content)
        workbook = load_workbook(content)

        # Verificar que el libro de trabajo tiene una hoja de cálculo activa
        self.assertTrue("Sheet" in workbook.sheetnames)

        # Obtener la hoja de cálculo activa
        sheet = workbook.active

        # Verificar que los encabezados están presentes
        self.assertEqual(sheet["A1"].value, "ID Votacion")
        self.assertEqual(sheet["B1"].value, "ID Votante")

        # Verificar que los datos están presentes
        self.assertEqual(sheet["A2"].value, self.census.voting_id)
        self.assertEqual(sheet["B2"].value, self.census.voter_id)


class AdminReuseCensusActionTest(TestCase):
    def setUp(self):
        self.site = AdminSite()
        self.model_admin = CensusAdmin(model=Census, admin_site=self.site)

    def test_reuse_action(self):
        census = Census.objects.create(voter_id=1, voting_id=1)

        request = HttpRequest()
        request.method = "POST"
        request.POST["id_to_reuse"] = "3"
        request._messages = messages.storage.default_storage(
            request
        )  # Necesario para testear código con mensajes

        self.model_admin.reuse_action(request, Census.objects.filter(pk=census.pk))

        # Llama de nuevo, para cubrir el código del caso donde Censo ya está presente en BD
        self.model_admin.reuse_action(request, Census.objects.filter(pk=census.pk))

        # Llama una última vez, para cubrir el código del caso donde Censo tenga ID nulo
        request.POST["id_to_reuse"] = None
        self.model_admin.reuse_action(request, Census.objects.filter(pk=census.pk))

        # Debe haber 2 censos con el mismo votante: El original y el creado reutilizando el previo
        self.assertEqual(len(Census.objects.filter(voter_id=census.voter_id)), 2)


class CensusByPreferenceTestCase(BaseTestCase):
    def setUp(self):
        super().setUp()
        self.census = CensusByPreference(voting_id=1, voter_id=1)
        self.census.save()

    def tearDown(self):
        super().tearDown()
        self.census = None

    def test_check_vote_permissions(self):
        response = self.client.get(
            "/census/bypreference/{}/?voter_id={}".format(1, 2), format="json"
        )
        self.assertEqual(response.status_code, 401)
        self.assertEqual(response.json(), "Invalid voter")

        response = self.client.get(
            "/census/bypreference/{}/?voter_id={}".format(1, 1), format="json"
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json(), "Valid voter")

    def test_list_voting(self):
        response = self.client.get(
            "/census/bypreference/?voting_id={}".format(1), format="json"
        )
        self.assertEqual(response.status_code, 401)

        self.login(user="noadmin")
        response = self.client.get(
            "/census/bypreference/?voting_id={}".format(1), format="json"
        )
        self.assertEqual(response.status_code, 403)

        self.login()
        response = self.client.get(
            "/census/bypreference/?voting_id={}".format(1), format="json"
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json(), {"voters": [1]})

    def test_add_new_voters_conflict(self):
        data = {"voting_id": 1, "voters": [1]}
        response = self.client.post("/census/bypreference/", data, format="json")
        self.assertEqual(response.status_code, 401)

        self.login(user="noadmin")
        response = self.client.post("/census/bypreference/", data, format="json")
        self.assertEqual(response.status_code, 403)

        self.login()
        response = self.client.post("/census/bypreference/", data, format="json")
        self.assertEqual(response.status_code, 409)

    def test_add_new_voters(self):
        data = {"voting_id": 2, "voters": [1, 2, 3, 4]}
        response = self.client.post("/census/bypreference/", data, format="json")
        self.assertEqual(response.status_code, 401)

        self.login(user="noadmin")
        response = self.client.post("/census/bypreference/", data, format="json")
        self.assertEqual(response.status_code, 403)

        self.login()
        response = self.client.post("/census/bypreference/", data, format="json")
        self.assertEqual(response.status_code, 201)
        self.assertEqual(
            len(data.get("voters")), CensusByPreference.objects.count() - 1
        )

    def test_destroy_voter(self):
        data = {"voters": [1], "voting_id": 1}
        response = self.client.delete(
            "/census/bypreference/{}/".format(1), data, format="json"
        )
        self.assertEqual(response.status_code, 204)
        self.assertEqual(0, CensusByPreference.objects.count())


class CensusByPreferenceTest(StaticLiveServerTestCase):
    def setUp(self):
        # Load base test functionality for decide
        self.base = BaseTestCase()
        self.base.setUp()

        options = webdriver.ChromeOptions()
        options.headless = True
        self.driver = webdriver.Chrome(options=options)

        super().setUp()

    def tearDown(self):
        super().tearDown()
        self.driver.quit()

        self.base.tearDown()

    def createCensusSuccess(self):
        self.cleaner.get(self.live_server_url + "/admin/login/?next=/admin/")
        self.cleaner.set_window_size(1280, 720)

        self.cleaner.find_element(By.ID, "id_username").click()
        self.cleaner.find_element(By.ID, "id_username").send_keys("decide")

        self.cleaner.find_element(By.ID, "id_password").click()
        self.cleaner.find_element(By.ID, "id_password").send_keys("decide")

        self.cleaner.find_element(By.ID, "id_password").send_keys("Keys.ENTER")

        self.cleaner.get(self.live_server_url + "/admin/censusbypreference/census/add")
        now = datetime.now()
        self.cleaner.find_element(By.ID, "id_voting_id").click()
        self.cleaner.find_element(By.ID, "id_voting_id").send_keys(
            now.strftime("%m%d%M%S")
        )
        self.cleaner.find_element(By.ID, "id_voter_id").click()
        self.cleaner.find_element(By.ID, "id_voter_id").send_keys(
            now.strftime("%m%d%M%S")
        )
        self.cleaner.find_element(By.NAME, "_save").click()

        self.assertTrue(
            self.cleaner.current_url
            == self.live_server_url + "/admin/census/censusbypreference"
        )

    def createCensusEmptyError(self):
        self.cleaner.get(self.live_server_url + "/admin/login/?next=/admin/")
        self.cleaner.set_window_size(1280, 720)

        self.cleaner.find_element(By.ID, "id_username").click()
        self.cleaner.find_element(By.ID, "id_username").send_keys("decide")

        self.cleaner.find_element(By.ID, "id_password").click()
        self.cleaner.find_element(By.ID, "id_password").send_keys("decide")

        self.cleaner.find_element(By.ID, "id_password").send_keys("Keys.ENTER")

        self.cleaner.get(self.live_server_url + "/admin/census/censusbypreference/add")

        self.cleaner.find_element(By.NAME, "_save").click()

        self.assertTrue(
            self.cleaner.find_element_by_xpath(
                "/html/body/div/div[3]/div/div[1]/div/form/div/p"
            ).text
            == "Please correct the errors below."
        )
        self.assertTrue(
            self.cleaner.current_url
            == self.live_server_url + "/admin/census/censusbypreference/add"
        )

    def createCensusValueError(self):
        self.cleaner.get(self.live_server_url + "/admin/login/?next=/admin/")
        self.cleaner.set_window_size(1280, 720)

        self.cleaner.find_element(By.ID, "id_username").click()
        self.cleaner.find_element(By.ID, "id_username").send_keys("decide")

        self.cleaner.find_element(By.ID, "id_password").click()
        self.cleaner.find_element(By.ID, "id_password").send_keys("decide")

        self.cleaner.find_element(By.ID, "id_password").send_keys("Keys.ENTER")

        self.cleaner.get(self.live_server_url + "/admin/census/censusbypreference/add")
        now = datetime.now()
        self.cleaner.find_element(By.ID, "id_voting_id").click()
        self.cleaner.find_element(By.ID, "id_voting_id").send_keys("64654654654654")
        self.cleaner.find_element(By.ID, "id_voter_id").click()
        self.cleaner.find_element(By.ID, "id_voter_id").send_keys("64654654654654")
        self.cleaner.find_element(By.NAME, "_save").click()

        self.assertTrue(
            self.cleaner.find_element_by_xpath(
                "/html/body/div/div[3]/div/div[1]/div/form/div/p"
            ).text
            == "Please correct the errors below."
        )
        self.assertTrue(
            self.cleaner.current_url
            == self.live_server_url + "/admin/census/censusbypreference/add"
        )


class CensusByPreferenceImportViewTest(BaseTestCase):
    def setUp(self):
        super().setUp()

    def create_voting(self):
        q = QuestionByPreference(desc="test_question")
        q.save()

        options = [
            QuestionOptionByPreference(
                question=q,
                option=f"option {i + 1}",
                preference=secrets.randbelow(10) + 1,
            )
            for i in range(3)
        ]
        QuestionOptionByPreference.objects.bulk_create(options)

        v = VotingByPreference(name="test_voting", question=q)
        v.save()

        auth, _ = Auth.objects.get_or_create(
            url=settings.BASEURL, defaults={"me": True, "name": "test_auth"}
        )
        v.auths.add(auth)

        return v

    def test_census_import_view(self):
        self.create_voting()
        test_group = "Test Group"

        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Voting ID", "Voter ID", "Group"])
        sheet.append([1, 1, test_group])
        sheet.append([1, 2, test_group])
        sheet.append([1, 1, test_group])  # Censo repetido, dará error

        file_buffer = BytesIO()
        workbook.save(file_buffer)
        file_buffer.seek(0)

        excel_file = SimpleUploadedFile("census.xlsx", file_buffer.read())

        url = reverse("import_census_by_preference")

        response = self.client.post(url, {"census_file": excel_file}, follow=True)

        self.assertEqual(response.status_code, 200)

        census_data = Census.objects.all()
        self.assertEqual(census_data.count(), 2)
        self.assertEqual(census_data[0].voting_id, 1)
        self.assertEqual(census_data[0].voter_id, 1)
        self.assertEqual(census_data[1].voting_id, 1)
        self.assertEqual(census_data[1].voter_id, 2)

        messages = list(response.context["messages"])
        expected_messages = [
            "Ya existe un registro para la pareja de voting_id=1, voter_id=1 y group=Test Group",
            "Importación finalizada",
        ]
        self.assertEqual([str(msg) for msg in messages], expected_messages)


class ByPreferenceAdminExportToExcelTest(TestCase):
    def setUp(self):
        self.site = AdminSite()
        self.model_admin = CensusByPreferenceAdmin(
            model=CensusByPreference, admin_site=self.site
        )
        self.census = CensusByPreference.objects.create(voter_id=1, voting_id=1)

    def test_export_to_excel(self):
        request = HttpRequest()
        request.method = "POST"

        # Simular la selección de objetos en el panel de administración
        queryset = CensusByPreference.objects.filter(pk=self.census.pk)

        # Ejecutar la acción de exportar a Excel
        response = self.model_admin.exportar_a_excel(request, queryset)

        # Verificar que la respuesta es un archivo Excel
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertTrue(
            "attachment; filename=exportacion_censo.xlsx"
            in response["Content-Disposition"]
        )

        # Cargar el libro de trabajo desde el contenido de la respuesta
        content = BytesIO(response.content)
        workbook = load_workbook(content)

        # Verificar que el libro de trabajo tiene una hoja de cálculo activa
        self.assertTrue("Sheet" in workbook.sheetnames)

        # Obtener la hoja de cálculo activa
        sheet = workbook.active

        # Verificar que los encabezados están presentes
        self.assertEqual(sheet["A1"].value, "ID Votacion")
        self.assertEqual(sheet["B1"].value, "ID Votante")

        # Verificar que los datos están presentes
        self.assertEqual(sheet["A2"].value, self.census.voting_id)
        self.assertEqual(sheet["B2"].value, self.census.voter_id)


class ByPreferenceAdminReuseCensusActionTest(TestCase):
    def setUp(self):
        self.site = AdminSite()
        self.model_admin = CensusByPreferenceAdmin(
            model=CensusByPreference, admin_site=self.site
        )

    def test_reuse_action(self):
        census = CensusByPreference.objects.create(voter_id=1, voting_id=1)

        request = HttpRequest()
        request.method = "POST"
        request.POST["id_to_reuse"] = "3"
        request._messages = messages.storage.default_storage(
            request
        )  # Necesario para testear código con mensajes

        self.model_admin.reuse_action(
            request, CensusByPreference.objects.filter(pk=census.pk)
        )

        # Llama de nuevo, para cubrir el código del caso donde Censo ya está presente en BD
        self.model_admin.reuse_action(
            request, CensusByPreference.objects.filter(pk=census.pk)
        )

        # Llama una última vez, para cubrir el código del caso donde Censo tenga ID nulo
        request.POST["id_to_reuse"] = None
        self.model_admin.reuse_action(
            request, CensusByPreference.objects.filter(pk=census.pk)
        )

        # Debe haber 2 censos con el mismo votante: El original y el creado reutilizando el previo
        self.assertEqual(
            len(CensusByPreference.objects.filter(voter_id=census.voter_id)), 2
        )


class CensusYesNoTestCase(BaseTestCase):
    def setUp(self):
        super().setUp()
        self.census = CensusYesNo(voting_id=1, voter_id=1)
        self.census.save()

    def tearDown(self):
        super().tearDown()
        self.census = None

    def test_check_vote_permissions(self):
        response = self.client.get(
            "/census/yesno/{}/?voter_id={}".format(1, 2), format="json"
        )
        self.assertEqual(response.status_code, 401)
        self.assertEqual(response.json(), "Invalid voter")

        response = self.client.get(
            "/census/yesno/{}/?voter_id={}".format(1, 1), format="json"
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json(), "Valid voter")

    def test_list_voting(self):
        response = self.client.get(
            "/census/yesno/?voting_id={}".format(1), format="json"
        )
        self.assertEqual(response.status_code, 401)

        self.login(user="noadmin")
        response = self.client.get(
            "/census/yesno/?voting_id={}".format(1), format="json"
        )
        self.assertEqual(response.status_code, 403)

        self.login()
        response = self.client.get(
            "/census/yesno/?voting_id={}".format(1), format="json"
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json(), {"voters": [1]})

    def test_add_new_voters_conflict(self):
        data = {"voting_id": 1, "voters": [1]}
        response = self.client.post("/census/yesno/", data, format="json")
        self.assertEqual(response.status_code, 401)

        self.login(user="noadmin")
        response = self.client.post("/census/yesno/", data, format="json")
        self.assertEqual(response.status_code, 403)

        self.login()
        response = self.client.post("/census/yesno/", data, format="json")
        self.assertEqual(response.status_code, 409)

    def test_add_new_voters(self):
        data = {"voting_id": 2, "voters": [1, 2, 3, 4]}
        response = self.client.post("/census/yesno/", data, format="json")
        self.assertEqual(response.status_code, 401)

        self.login(user="noadmin")
        response = self.client.post("/census/yesno/", data, format="json")
        self.assertEqual(response.status_code, 403)

        self.login()
        response = self.client.post("/census/yesno/", data, format="json")
        self.assertEqual(response.status_code, 201)
        self.assertEqual(len(data.get("voters")), CensusYesNo.objects.count() - 1)

    def test_destroy_voter(self):
        data = {"voters": [1], "voting_id": 1}
        response = self.client.delete(
            "/census/yesno/{}/".format(1), data, format="json"
        )
        self.assertEqual(response.status_code, 204)
        self.assertEqual(0, CensusYesNo.objects.count())


class CensusYesNoTest(StaticLiveServerTestCase):
    def setUp(self):
        # Load base test functionality for decide
        self.base = BaseTestCase()
        self.base.setUp()

        options = webdriver.ChromeOptions()
        options.headless = True
        self.driver = webdriver.Chrome(options=options)

        super().setUp()

    def tearDown(self):
        super().tearDown()
        self.driver.quit()

        self.base.tearDown()

    def createCensusSuccess(self):
        self.cleaner.get(self.live_server_url + "/admin/login/?next=/admin/")
        self.cleaner.set_window_size(1280, 720)

        self.cleaner.find_element(By.ID, "id_username").click()
        self.cleaner.find_element(By.ID, "id_username").send_keys("decide")

        self.cleaner.find_element(By.ID, "id_password").click()
        self.cleaner.find_element(By.ID, "id_password").send_keys("decide")

        self.cleaner.find_element(By.ID, "id_password").send_keys("Keys.ENTER")

        self.cleaner.get(self.live_server_url + "/admin/census/censusyesno/add/")
        now = datetime.now()
        self.cleaner.find_element(By.ID, "id_voting_id").click()
        self.cleaner.find_element(By.ID, "id_voting_id").send_keys(
            now.strftime("%m%d%M%S")
        )
        self.cleaner.find_element(By.ID, "id_voter_id").click()
        self.cleaner.find_element(By.ID, "id_voter_id").send_keys(
            now.strftime("%m%d%M%S")
        )
        self.cleaner.find_element(By.ID, "id_group").click()
        self.cleaner.find_element(By.ID, "id_group").send_keys(now.strftime("%m%d%M%S"))
        self.cleaner.find_element(By.NAME, "_save").click()

        self.assertTrue(
            self.cleaner.current_url
            == self.live_server_url + "/admin/census/censusyesno/"
        )

    def createCensusEmptyError(self):
        self.cleaner.get(self.live_server_url + "/admin/login/?next=/admin/")
        self.cleaner.set_window_size(1280, 720)

        self.cleaner.find_element(By.ID, "id_username").click()
        self.cleaner.find_element(By.ID, "id_username").send_keys("decide")

        self.cleaner.find_element(By.ID, "id_password").click()
        self.cleaner.find_element(By.ID, "id_password").send_keys("decide")

        self.cleaner.find_element(By.ID, "id_password").send_keys("Keys.ENTER")

        self.cleaner.get(self.live_server_url + "/admin/census/censusyesno/add/")

        self.cleaner.find_element(By.NAME, "_save").click()

        self.assertTrue(
            self.cleaner.find_element_by_xpath(
                "/html/body/div/div[3]/div/div[1]/div/form/div/p"
            ).text
            == "Please correct the errors below."
        )
        self.assertTrue(
            self.cleaner.current_url
            == self.live_server_url + "/admin/census/censusyesno/add/"
        )

    def createCensusValueError(self):
        self.cleaner.get(self.live_server_url + "/admin/login/?next=/admin/")
        self.cleaner.set_window_size(1280, 720)

        self.cleaner.find_element(By.ID, "id_username").click()
        self.cleaner.find_element(By.ID, "id_username").send_keys("decide")

        self.cleaner.find_element(By.ID, "id_password").click()
        self.cleaner.find_element(By.ID, "id_password").send_keys("decide")

        self.cleaner.find_element(By.ID, "id_password").send_keys("Keys.ENTER")

        self.cleaner.get(self.live_server_url + "/admin/census/censusyesno/add")
        now = datetime.now()
        self.cleaner.find_element(By.ID, "id_voting_id").click()
        self.cleaner.find_element(By.ID, "id_voting_id").send_keys("64654654654654")
        self.cleaner.find_element(By.ID, "id_voter_id").click()
        self.cleaner.find_element(By.ID, "id_voter_id").send_keys("64654654654654")
        self.cleaner.find_element(By.ID, "id_group").click()
        self.cleaner.find_element(By.ID, "id_group").send_keys("64654654654654")
        self.cleaner.find_element(By.NAME, "_save").click()

        self.assertTrue(
            self.cleaner.find_element_by_xpath(
                "/html/body/div/div[3]/div/div[1]/div/form/div/p"
            ).text
            == "Please correct the errors below."
        )
        self.assertTrue(
            self.cleaner.current_url
            == self.live_server_url + "/admin/census/censusyesno/add"
        )


class CensusYesNoImportViewTest(BaseTestCase):
    def setUp(self):
        super().setUp()

    def create_voting_yesno(self):
        q = QuestionYesNo(desc="test_question")
        q.save()

        v = VotingYesNo(name="test_voting", question=q)
        v.save()

        auth, _ = Auth.objects.get_or_create(
            url=settings.BASEURL, defaults={"me": True, "name": "test_auth"}
        )
        v.auths.add(auth)

        return v

    def test_census_import_view(self):
        self.create_voting_yesno()

        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Voting Yesno ID", "Voter ID"])
        sheet.append([1, 1])
        sheet.append([1, 2])
        sheet.append([1, 1])  # Censo repetido, dará error

        file_buffer = BytesIO()
        workbook.save(file_buffer)
        file_buffer.seek(0)

        excel_file = SimpleUploadedFile("census.xlsx", file_buffer.read())

        url = reverse("import_census_yesno")

        response = self.client.post(url, {"census_file": excel_file}, follow=True)

        self.assertEqual(response.status_code, 200)

        census_data = CensusYesNo.objects.all()
        self.assertEqual(census_data.count(), 2)
        self.assertEqual(census_data[0].voting_id, 1)
        self.assertEqual(census_data[0].voter_id, 1)
        self.assertEqual(census_data[1].voting_id, 1)
        self.assertEqual(census_data[1].voter_id, 2)

        messages = list(response.context["messages"])
        expected_messages = [
            "Ya existe un registro para la pareja de voting_id=1 y voter_id=1",
            "Importación finalizada",
        ]
        self.assertEqual([str(msg) for msg in messages], expected_messages)


class YesNoAdminExportToExcelTest(TestCase):
    def setUp(self):
        self.site = AdminSite()
        self.model_admin = CensusYesNoAdmin(model=CensusYesNo, admin_site=self.site)
        self.census = CensusYesNo.objects.create(voter_id=1, voting_id=1)

    def test_export_to_excel(self):
        request = HttpRequest()
        request.method = "POST"

        # Simular la selección de objetos en el panel de administración
        queryset = CensusYesNo.objects.filter(pk=self.census.pk)

        # Ejecutar la acción de exportar a Excel
        response = self.model_admin.exportar_a_excel(request, queryset)

        # Verificar que la respuesta es un archivo Excel
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertTrue(
            "attachment; filename=exportacion_censo.xlsx"
            in response["Content-Disposition"]
        )

        # Cargar el libro de trabajo desde el contenido de la respuesta
        content = BytesIO(response.content)
        workbook = load_workbook(content)

        # Verificar que el libro de trabajo tiene una hoja de cálculo activa
        self.assertTrue("Sheet" in workbook.sheetnames)

        # Obtener la hoja de cálculo activa
        sheet = workbook.active

        # Verificar que los encabezados están presentes
        self.assertEqual(sheet["A1"].value, "ID Votacion")
        self.assertEqual(sheet["B1"].value, "ID Votante")

        # Verificar que los datos están presentes
        self.assertEqual(sheet["A2"].value, self.census.voting_id)
        self.assertEqual(sheet["B2"].value, self.census.voter_id)


class YesNoAdminReuseCensusActionTest(TestCase):
    def setUp(self):
        self.site = AdminSite()
        self.model_admin = CensusYesNoAdmin(model=CensusYesNo, admin_site=self.site)

    def test_reuse_action(self):
        census = CensusYesNo.objects.create(voter_id=1, voting_id=1)

        request = HttpRequest()
        request.method = "POST"
        request.POST["id_to_reuse"] = "3"
        request._messages = messages.storage.default_storage(
            request
        )  # Necesario para testear código con mensajes

        self.model_admin.reuse_action(request, CensusYesNo.objects.filter(pk=census.pk))

        # Llama de nuevo, para cubrir el código del caso donde Censo ya está presente en BD
        self.model_admin.reuse_action(request, CensusYesNo.objects.filter(pk=census.pk))

        # Llama una última vez, para cubrir el código del caso donde Censo tenga ID nulo
        request.POST["id_to_reuse"] = None
        self.model_admin.reuse_action(request, CensusYesNo.objects.filter(pk=census.pk))

        # Debe haber 2 censos con el mismo votante: El original y el creado reutilizando el previo
        self.assertEqual(len(CensusYesNo.objects.filter(voter_id=census.voter_id)), 2)

class CensusMultiChoiceTestCase(BaseTestCase):
    def setUp(self):
        super().setUp()
        self.census = CensusMultiChoice(voting_id=1, voter_id=1)
        self.census.save()

    def tearDown(self):
        super().tearDown()
        self.census = None

    def test_check_vote_permissions(self):
        response = self.client.get(
            "/census/multichoice/{}/?voter_id={}".format(1, 2), format="json"
        )
        self.assertEqual(response.status_code, 401)
        self.assertEqual(response.json(), "Invalid voter")

        response = self.client.get(
            "/census/multichoice/{}/?voter_id={}".format(1, 1), format="json"
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json(), "Valid voter")

    def test_list_voting(self):
        response = self.client.get(
            "/census/multichoice/?voting_id={}".format(1), format="json"
        )
        self.assertEqual(response.status_code, 401)

        self.login(user="noadmin")
        response = self.client.get(
            "/census/multichoice/?voting_id={}".format(1), format="json"
        )
        self.assertEqual(response.status_code, 403)

        self.login()
        response = self.client.get(
            "/census/multichoice/?voting_id={}".format(1), format="json"
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json(), {"voters": [1]})

    def test_add_new_voters_conflict(self):
        data = {"voting_id": 1, "voters": [1]}
        response = self.client.post("/census/multichoice/", data, format="json")
        self.assertEqual(response.status_code, 401)

        self.login(user="noadmin")
        response = self.client.post("/census/multichoice/", data, format="json")
        self.assertEqual(response.status_code, 403)

        self.login()
        response = self.client.post("/census/multichoice/", data, format="json")
        self.assertEqual(response.status_code, 409)

    def test_add_new_voters(self):
        data = {"voting_id": 2, "voters": [1, 2, 3, 4]}
        response = self.client.post("/census/multichoice/", data, format="json")
        self.assertEqual(response.status_code, 401)

        self.login(user="noadmin")
        response = self.client.post("/census/multichoice/", data, format="json")
        self.assertEqual(response.status_code, 403)

        self.login()
        response = self.client.post("/census/multichoice/", data, format="json")
        self.assertEqual(response.status_code, 201)
        self.assertEqual(
            len(data.get("voters")), CensusMultiChoice.objects.count() - 1
        )

    def test_destroy_voter(self):
        data = {"voters": [1], "voting_id": 1}
        response = self.client.delete(
            "/census/multichoice/{}/".format(1), data, format="json"
        )
        self.assertEqual(response.status_code, 204)
        self.assertEqual(0, CensusMultiChoice.objects.count())


class CensusMultiChoiceTest(StaticLiveServerTestCase):
    def setUp(self):
        # Load base test functionality for decide
        self.base = BaseTestCase()
        self.base.setUp()

        options = webdriver.ChromeOptions()
        options.headless = True
        self.driver = webdriver.Chrome(options=options)

        super().setUp()

    def tearDown(self):
        super().tearDown()
        self.driver.quit()

        self.base.tearDown()

    def createCensusSuccess(self):
        self.cleaner.get(self.live_server_url + "/admin/login/?next=/admin/")
        self.cleaner.set_window_size(1280, 720)

        self.cleaner.find_element(By.ID, "id_username").click()
        self.cleaner.find_element(By.ID, "id_username").send_keys("decide")

        self.cleaner.find_element(By.ID, "id_password").click()
        self.cleaner.find_element(By.ID, "id_password").send_keys("decide")

        self.cleaner.find_element(By.ID, "id_password").send_keys("Keys.ENTER")

        self.cleaner.get(self.live_server_url + "/admin/censusmultichoice/census/add")
        now = datetime.now()
        self.cleaner.find_element(By.ID, "id_voting_id").click()
        self.cleaner.find_element(By.ID, "id_voting_id").send_keys(
            now.strftime("%m%d%M%S")
        )
        self.cleaner.find_element(By.ID, "id_voter_id").click()
        self.cleaner.find_element(By.ID, "id_voter_id").send_keys(
            now.strftime("%m%d%M%S")
        )
        self.cleaner.find_element(By.NAME, "_save").click()

        self.assertTrue(
            self.cleaner.current_url
            == self.live_server_url + "/admin/census/censusmultichoice"
        )

    def createCensusEmptyError(self):
        self.cleaner.get(self.live_server_url + "/admin/login/?next=/admin/")
        self.cleaner.set_window_size(1280, 720)

        self.cleaner.find_element(By.ID, "id_username").click()
        self.cleaner.find_element(By.ID, "id_username").send_keys("decide")

        self.cleaner.find_element(By.ID, "id_password").click()
        self.cleaner.find_element(By.ID, "id_password").send_keys("decide")

        self.cleaner.find_element(By.ID, "id_password").send_keys("Keys.ENTER")

        self.cleaner.get(self.live_server_url + "/admin/census/censusmultichoice/add")

        self.cleaner.find_element(By.NAME, "_save").click()

        self.assertTrue(
            self.cleaner.find_element_by_xpath(
                "/html/body/div/div[3]/div/div[1]/div/form/div/p"
            ).text
            == "Please correct the errors below."
        )
        self.assertTrue(
            self.cleaner.current_url
            == self.live_server_url + "/admin/census/censusmultichoice/add"
        )

    def createCensusValueError(self):
        self.cleaner.get(self.live_server_url + "/admin/login/?next=/admin/")
        self.cleaner.set_window_size(1280, 720)

        self.cleaner.find_element(By.ID, "id_username").click()
        self.cleaner.find_element(By.ID, "id_username").send_keys("decide")

        self.cleaner.find_element(By.ID, "id_password").click()
        self.cleaner.find_element(By.ID, "id_password").send_keys("decide")

        self.cleaner.find_element(By.ID, "id_password").send_keys("Keys.ENTER")

        self.cleaner.get(self.live_server_url + "/admin/census/censusmultichoice/add")
        now = datetime.now()
        self.cleaner.find_element(By.ID, "id_voting_id").click()
        self.cleaner.find_element(By.ID, "id_voting_id").send_keys("64654654654654")
        self.cleaner.find_element(By.ID, "id_voter_id").click()
        self.cleaner.find_element(By.ID, "id_voter_id").send_keys("64654654654654")
        self.cleaner.find_element(By.NAME, "_save").click()

        self.assertTrue(
            self.cleaner.find_element_by_xpath(
                "/html/body/div/div[3]/div/div[1]/div/form/div/p"
            ).text
            == "Please correct the errors below."
        )
        self.assertTrue(
            self.cleaner.current_url
            == self.live_server_url + "/admin/census/censusmultichoice/add"
        )


class CensusMultiChoiceImportViewTest(BaseTestCase):
    def setUp(self):
        super().setUp()

    def create_voting_multichoice(self):
        q = QuestionMultiChoice(desc="test_question")
        q.save()

        options = [
            QuestionOptionMultiChoice(
                question=q,
                option=f"option {i + 1}",
                multichoice=secrets.randbelow(10) + 1,
            )
            for i in range(3)
        ]
        QuestionOptionMultiChoice.objects.bulk_create(options)

        v = VotingMultiChoice(name="test_voting", question=q)
        v.save()

        auth, _ = Auth.objects.get_or_create(
            url=settings.BASEURL, defaults={"me": True, "name": "test_auth"}
        )
        v.auths.add(auth)

        return v

    def test_census_import_view(self):
        self.create_voting_multichoice()
        test_group = "Test Group"

        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Voting ID", "Voter ID", "Group"])
        sheet.append([1, 1, test_group])
        sheet.append([1, 2, test_group])
        sheet.append([1, 1, test_group])  # Censo repetido, dará error

        file_buffer = BytesIO()
        workbook.save(file_buffer)
        file_buffer.seek(0)

        excel_file = SimpleUploadedFile("census.xlsx", file_buffer.read())

        url = reverse("import_census_multichoice")

        response = self.client.post(url, {"census_file": excel_file}, follow=True)

        self.assertEqual(response.status_code, 200)

        census_data = Census.objects.all()
        self.assertEqual(census_data.count(), 2)
        self.assertEqual(census_data[0].voting_id, 1)
        self.assertEqual(census_data[0].voter_id, 1)
        self.assertEqual(census_data[1].voting_id, 1)
        self.assertEqual(census_data[1].voter_id, 2)

        messages = list(response.context["messages"])
        expected_messages = [
            "Ya existe un registro para la pareja de voting_id=1, voter_id=1 y group=Test Group",
            "Importación finalizada",
        ]
        self.assertEqual([str(msg) for msg in messages], expected_messages)


class MultiChoiceAdminExportToExcelTest(TestCase):
    def setUp(self):
        self.site = AdminSite()
        self.model_admin = CensusMultiChoiceAdmin(
            model=CensusMultiChoice, admin_site=self.site
        )
        self.census = CensusMultiChoice.objects.create(voter_id=1, voting_id=1)

    def test_export_to_excel(self):
        request = HttpRequest()
        request.method = "POST"

        # Simular la selección de objetos en el panel de administración
        queryset = CensusMultiChoice.objects.filter(pk=self.census.pk)

        # Ejecutar la acción de exportar a Excel
        response = self.model_admin.exportar_a_excel(request, queryset)

        # Verificar que la respuesta es un archivo Excel
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertTrue(
            "attachment; filename=exportacion_censo.xlsx"
            in response["Content-Disposition"]
        )

        # Cargar el libro de trabajo desde el contenido de la respuesta
        content = BytesIO(response.content)
        workbook = load_workbook(content)

        # Verificar que el libro de trabajo tiene una hoja de cálculo activa
        self.assertTrue("Sheet" in workbook.sheetnames)

        # Obtener la hoja de cálculo activa
        sheet = workbook.active

        # Verificar que los encabezados están presentes
        self.assertEqual(sheet["A1"].value, "ID Votacion")
        self.assertEqual(sheet["B1"].value, "ID Votante")

        # Verificar que los datos están presentes
        self.assertEqual(sheet["A2"].value, self.census.voting_id)
        self.assertEqual(sheet["B2"].value, self.census.voter_id)


class MultiChoiceAdminReuseCensusActionTest(TestCase):
    def setUp(self):
        self.site = AdminSite()
        self.model_admin = CensusMultiChoiceAdmin(
            model=CensusMultiChoice, admin_site=self.site
        )

    def test_reuse_action(self):
        census = CensusMultiChoice.objects.create(voter_id=1, voting_id=1)

        request = HttpRequest()
        request.method = "POST"
        request.POST["id_to_reuse"] = "3"
        request._messages = messages.storage.default_storage(
            request
        )  # Necesario para testear código con mensajes

        self.model_admin.reuse_action(
            request, CensusMultiChoice.objects.filter(pk=census.pk)
        )

        # Llama de nuevo, para cubrir el código del caso donde Censo ya está presente en BD
        self.model_admin.reuse_action(
            request, CensusMultiChoice.objects.filter(pk=census.pk)
        )

        # Llama una última vez, para cubrir el código del caso donde Censo tenga ID nulo
        request.POST["id_to_reuse"] = None
        self.model_admin.reuse_action(
            request, CensusMultiChoice.objects.filter(pk=census.pk)
        )

        # Debe haber 2 censos con el mismo votante: El original y el creado reutilizando el previo
        self.assertEqual(
            len(CensusMultiChoice.objects.filter(voter_id=census.voter_id)), 2
        )